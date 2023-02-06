import openpyxl
import pytest
from crm.agency_cabinet.ord.server.src.celery.tasks.transfer.export_reports import AgencyDirectReportExporter
from crm.agency_cabinet.ord.server.src.config import MDS_SETTINGS, REPORTS_MDS_SETTINGS


@pytest.fixture
def agency_direct_report_exporter(mds_cfg, fixture_celery_report_export):
    return AgencyDirectReportExporter(
        MDS_SETTINGS,
        REPORTS_MDS_SETTINGS,
        fixture_celery_report_export[0].id,
        fixture_celery_report_export[0].report_id
    )


async def test_export_report(
    agency_direct_report_exporter,
    fixture_celery_report_export,
    fixture_campaign,
    fixture_celery_client_rows,
):
    agency_direct_report_exporter._report_id = fixture_celery_report_export[0].report_id
    report_file = await agency_direct_report_exporter._make_report()
    workbook = openpyxl.load_workbook(report_file.name)
    worksheet = workbook.active

    expected_common_headers = [
        'Логин/Кампания',
        'Площадка',
        'Я-Агентство',
        'Мой контрагент',
        'Договор между агентством и его контрагентом',
        'Контрагент конечного рекламодателя',
        'Конечный рекламодатель',
        'Договор между конечным рекламодателем и его исполнителем'
    ]

    common_headers = [
        worksheet.cell(row=1, column=i).value for i in range(1, worksheet.max_column)
            if worksheet.cell(row=1, column=i).value is not None
    ]
    assert common_headers == expected_common_headers

    expected_headers = [
        # Логин/Кампания
        'Client ID',
        'Login',
        'Рекламная кампания',
        'Рекомендуемая сумма (с НДС)',
        'Идентификатор креатива',
        # Площадка
        'Акт',
        'Сумма (с НДС)',
        'Признак "с НДС"',
        # Я-Агентство
        'ИНН',
        'Тип организации',
        'ОПФ и полное наименование',
        'Абонентский номер мобильного телефона',
        'Номер электронного средства платежа',
        'Регистрационный номер либо его аналог',
        'Номер налогоплательщика либо его аналог в стране регистрации',
        'Код страны регистрации юрлица в соответствии с ОКСМ',
        # Мой контрагент
        'ИНН',
        'Тип организации',
        'ОПФ и полное наименование',
        'Абонентский номер мобильного телефона',
        'Номер электронного средства платежа',
        'Регистрационный номер либо его аналог',
        'Номер налогоплательщика либо его аналог в стране регистрации',
        'Код страны регистрации юрлица в соответствии с ОКСМ',
        # Договор между агентством и его контрагентом
        'Номер договора',
        'Тип договора',
        'Описание осуществляемых посредником-представителем действий',
        'Сведения о предмете договора',
        'Дата договора / доп.соглашения',
        'Цена договора',
        'Признак "с НДС"',
        'Акт',
        'Сумма (с НДС)',
        'Признак "с НДС"',
        'Мой контрагент является конечным',
        # Контрагент конечного рекламодателя'
        'ИНН',
        'Тип организации',
        'ОПФ и полное наименование',
        'Абонентский номер мобильного телефона',
        'Номер электронного средства платежа',
        'Регистрационный номер либо его аналог',
        'Номер налогоплательщика либо его аналог в стране регистрации',
        'Код страны регистрации юрлица в соответствии с ОКСМ',
        # Конечный рекламодатель
        'ИНН',
        'Тип организации',
        'ОПФ и полное наименование',
        'Абонентский номер мобильного телефона',
        'Номер электронного средства платежа',
        'Регистрационный номер либо его аналог',
        'Номер налогоплательщика либо его аналог в стране регистрации',
        'Код страны регистрации юрлица в соответствии с ОКСМ',
        # Договор между конечным рекламодателем и его исполнителем
        'Номер договора',
        'Тип договора',
        'Описание осуществляемых посредником-представителем действий',
        'Сведения о предмете договора',
        'Дата договора / доп.соглашения',
        'Цена договора',
        'Признак "с НДС"',
    ]

    headers = [worksheet.cell(row=2, column=i).value for i in range(1, worksheet.max_column + 1)]
    assert headers == expected_headers

    expected_row_client_1_all_campanies = [
        # Логин/Кампания
        'client_id_1', 'client_login_1', 'Все кампании'
    ]

    expected_row_client_1 = [
        # Логин/Кампания
        'client_id_1', 'client_login_1', 'campaign_name_1', 3000, 'campaign_eid_1',
        # Площадка
        'ad_distributor_act_eid_1', 3000, 'да',
        # Я-Агентство
        '123456789', 'ul', 'agency_name_1', None, None, None, None, None,
        # Мой контрагент
        '234567891', 'ul', 'contragent_name_1', None, None, None, None, None,
        # Договор между агентством и его контрагентом
        'partner_contract_eid_1', 'contract', 'distribution', 'distribution', '2022-05-01', 3000, 'да',
        'act1', 1000, 'да', False,
        # Контрагент конечного рекламодателя'
        '12121212', 'ul', 'advertiser_contragent_name_1', None, None, None, None, None,
        # Конечный рекламодатель
        '111222333', 'ul', 'advertiser_name_1', None, None, None, None, None,
        # Договор между конечным рекламодателем и его исполнителем
        'advertiser_contract_eid_1', 'contract', 'distribution', 'distribution', '2022-05-01', 3000, 'да'
    ]

    expected_row_client_1_2 = [
        # Логин/Кампания
        'client_id_1', 'client_login_1', 'campaign_name_1', 3000, 'campaign_eid_1',
        # Площадка
        'ad_distributor_act_eid_1', 3000, 'да',
        # Я-Агентство
        '123456789', 'ul', 'agency_name_1', None, None, None, None, None,
        # Мой контрагент
        '234567891', 'ul', 'contragent_name_1', None, None, None, None, None,
        # Договор между агентством и его контрагентом
        'partner_contract_eid_1', 'contract', 'distribution', 'distribution', '2022-05-01', 3000, 'да',
        'act2', 2000, 'да', False,
        # Контрагент конечного рекламодателя'
        '12121212', 'ul', 'advertiser_contragent_name_1', None, None, None, None, None,
        # Конечный рекламодатель
        '111222333', 'ul', 'advertiser_name_1', None, None, None, None, None,
        # Договор между конечным рекламодателем и его исполнителем
        'advertiser_contract_eid_1', 'contract', 'distribution', 'distribution', '2022-05-01', 3000, 'да'
    ]

    expected_row_client_2_all_campanies = [
        # Логин/Кампания
        'client_id_2', 'client_login_2', 'Все кампании'
    ]

    expected_row_client_2 = [
        # Логин/Кампания
        'client_id_2', 'client_login_2', 'campaign_name_2', 3000, 'campaign_eid_2',
        # Площадка
        'ad_distributor_act_eid_2', 3000, 'да',
        # Я-Агентство
        '123456789', 'ul', 'agency_name_1', None, None, None, None, None,
        # Мой контрагент
        '345678912', 'ul', 'contragent_name_2', None, None, None, None, None,
        # Договор между агентством и его контрагентом
        'partner_contract_eid_2', 'contract', 'distribution', 'distribution', '2022-05-01', 3000, 'да',
        'act3', 3000, 'да', False,
        # Контрагент конечного рекламодателя'
        '21212121', 'ul', 'advertiser_contragent_name_2', None, None, None, None, None,
        # Конечный рекламодатель
        '222333111', 'ul', 'advertiser_name_2', None, None, None, None, None,
        # Договор между конечным рекламодателем и его исполнителем
        'advertiser_contract_eid_2', 'contract', 'distribution', 'distribution', '2022-05-01', 3000, 'да'
    ]

    # client 1
    row_client_1_all_campanies = [worksheet.cell(row=3, column=i).value for i in range(1, 4)]
    assert row_client_1_all_campanies == expected_row_client_1_all_campanies

    row_client_1 = [worksheet.cell(row=4, column=i).value for i in range(1, worksheet.max_column + 1)]
    assert row_client_1 == expected_row_client_1

    row_client_1_2 = [worksheet.cell(row=5, column=i).value for i in range(1, worksheet.max_column + 1)]
    assert row_client_1_2 == expected_row_client_1_2

    # client 2
    row_client_2_all_campanies = [worksheet.cell(row=6, column=i).value for i in range(1, 4)]
    assert row_client_2_all_campanies == expected_row_client_2_all_campanies

    row_client_2 = [worksheet.cell(row=7, column=i).value for i in range(1, worksheet.max_column + 1)]
    assert row_client_2 == expected_row_client_2
